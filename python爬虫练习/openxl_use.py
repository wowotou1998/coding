from openpyxl import Workbook
from openpyxl import load_workbook
# 新建一个工作簿
wb = Workbook()
# 新建一个工作表
# ws = wb.active
# 创建新的表
ws1 = wb.create_sheet("总库",0)
ws2 = wb.create_sheet("用人单位",1)
ws3 = wb.create_sheet("参会代表",2)
ws4 = wb.create_sheet("需求信息",3)
list1 = ['企业名称', '时间', '地点',  '联系方式', '专业']
# 访问工作簿中ws1工作表中的细胞
# row是行 column是 列
ws1.cell(row=1, column=2, value=10)
# 对100*100的细胞都设置为12

for row in ws1.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
print(wb.sheetnames)
wb.save("E:\\exa_openpyxl_jyw1133.xlsx")

